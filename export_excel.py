from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pymysql

select_type = "select distinct question_type from exam_questions order by question_type asc;"
select_sql = "SELECT id, chapter_no, title, question_type, options, correct_answer, short_answer FROM exam_questions WHERE question_type = %s;"

conn = pymysql.connect(
    host="localhost",
    user="root",
    password="12345",
    database="user",
    port=3306,
    charset="utf8mb4"
)

cursor = conn.cursor()

wb = Workbook()

dan_xuan = wb.active
dan_xuan.title = "单选题"
daun_xuan_headers = ["章节标题", "题目", "选项", "答案(正确/错误)"]
pan_duan_headers = ["章节标题", "题目", "答案"]
dan_xuan.append(daun_xuan_headers)
duo_xuan = wb.create_sheet(title="多选题")
duo_xuan.append(daun_xuan_headers)
pan_duan = wb.create_sheet(title="判断题")
pan_duan.append(pan_duan_headers)


def export_excel(path):
    cursor.execute(select_type)
    ques_type = cursor.fetchall()

    if ques_type is None:
        print("未找到题目类型")
        return

    # 根据题目类型导出到excel
    for i, v in enumerate(ques_type):
        cursor.execute(select_sql, v)
        res = cursor.fetchall()

        if v[0] == 'dan_xuan':
            for k in res:
                temp = []
                # 章节题目
                temp.append(k[1])
                # 题目
                temp.append(k[2])
                # 选项
                options = ' '.join(k[4].strip().replace("\n", "").split("<|>"))
                temp.append(options)
                # 正常答案
                tag = "对" if k[5][-1] == "d" else "错"
                answer = k[5][:-1] + "("+tag+")"
                temp.append(answer)
                dan_xuan.append(temp)
        elif v[0] == 'duo_xuan':
            for k in res:
                temp = []
                # 章节题目
                temp.append(k[1])
                # 题目
                temp.append(k[2])
                # 选项
                options = ' '.join(k[4].strip().replace("\n", "").split("<|>"))
                temp.append(options)
                # 正常答案
                tag = "对" if k[5][-1] == "d" else "错"
                answer = k[5][:-1] + "("+tag+")"
                temp.append(answer)
                duo_xuan.append(temp)
        elif v[0] == 'pan_duan':
            for k in res:
                temp = []
                # 章节题目
                temp.append(k[1])
                # 题目
                temp.append(k[2])
                # 选项
                # options = ' '.join(k[3].strip().replace("\n", "").split("<|>"))
                # temp.append(options)
                # 正常答案
                if k[5][-1] == "x":
                    answer = "对" if k[5][:-1] == "错" else "错"
                else:
                    answer = k[5][:-1]
                temp.append(answer)
                pan_duan.append(temp)

    format(dan_xuan)
    format(duo_xuan)
    format(pan_duan)
    wb.save(path)


def format(ws):
    # 设置整个 sheet 的字体（遍历所有单元格）
    font = Font(name="微软雅黑", size=18)  # 字体：微软雅黑，11号
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font

    # 设置列宽自动化
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # 获取列字母
        for cell in col:
            if cell.value:  # 计算每个单元格内容长度
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 20  # 留点空隙

if __name__ == '__main__':
    export_excel("数字化转型练习题库(不全).xlsx")
